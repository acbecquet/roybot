# scripts/make_bom.py
"""One-shot generator for the staged Roybot Bill of Materials.

STAGE A = first prototype, buy now (drive + chase brain + GrowBot senses + portable power +
perception OFFLOADED to the PC). STAGE B and C build on A by swapping the perception backend
(see docs/superpowers/specs/2026-06-28-roybot-perception-architecture.md):
  B = standalone onboard colored-collar blob tracker (software-only on A's hardware + a collar)
  C = full standalone, no collar (brain upgrade: Pi 5 + AI accelerator)
Tools assumed on hand (becqu solders at work). Prices are rough ballparks, not quotes.

Run once:  .venv/Scripts/python scripts/make_bom.py
"""
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule

OUT = Path("docs/phase2/roybot-prototype-BOM.xlsx")
STATUSES = ["Buy", "Print", "Have", "Spare"]
STAGES = ["A", "B", "C"]

# (Stage, Status, Group, Component, Qty, Spec / note, Source, Est. unit USD)
ROWS = [
    # ================= STAGE A — first prototype (buy now) =================
    # -- Drivetrain (N20 + encoders: closed-loop, stall cutoff, odometry) --
    ("A", "Buy", "Drivetrain", "Pololu HPCB 6V 100:1 micro metal gearmotor, EXTENDED back-shaft", 2,
     "Verified pick (~0.75 m/s, strong torque). EXTENDED shaft REQUIRED for the encoder. Cheaper alt: generic GA12-N20 6V w/ integrated encoder (~$12 ea, 2-packs) saves ~$20 - Pololu chosen for spec consistency", "Pololu", 18),
    ("A", "Buy", "Drivetrain", "Pololu magnetic quadrature encoder pair (micro metal gearmotor)", 1,
     "1200 CPR/wheel-rev at 100:1 (already x4-decoded). VCC=3.3V. Closed-loop speed (matches sim), stall cutoff, odometry", "Pololu", 9),
    ("A", "Buy", "Drivetrain", "60 mm wheel for N20 (3 mm D-shaft), rubber tire", 2,
     "3216 wheels don't fit the N20 shaft. Rubber tire for grip", "Pololu / Amazon", 4),
    ("A", "Buy", "Drivetrain", "TB6612FNG dual H-bridge breakout", 1,
     "Motor driver. VM = 6V motor rail, VCC = 3.3V, PWM >=20 kHz. Closed-loop via encoders + firmware stall cutoff", "Adafruit / Pololu", 5),
    # -- Compute / sensing --
    ("A", "Buy", "Compute", "microSD card, 32 GB Class 10 / A1", 1,
     "NOT in your Pi kit. Raspberry Pi OS Lite 64-bit (GrowBot #2)", "Amazon", 8),
    ("A", "Buy", "Sensing", "OV5647 camera - Pi ZERO version (narrow 22->15 CSI ribbon)", 1,
     "HIGH PRIORITY. MUST be the Pi Zero narrow-ribbon version (GrowBot #7). Used by all 3 perception stages", "Amazon", 12),
    ("A", "Buy", "Sensing", "MPU-6050 IMU (GY-521)", 1,
     "Tip + no-motion stall reflexes; proprioception. I2C 0x68, 3.3V (GrowBot #6)", "Amazon", 4),
    # -- Voice & lights (GrowBot guts; high priority per becqu) --
    ("A", "Buy", "Voice/Lights", "INMP441 I2S microphone", 1, "HIGH PRIORITY. Wake-word / voice in (GrowBot #8)", "Amazon", 4),
    ("A", "Buy", "Voice/Lights", "MAX98357A I2S amplifier", 1, "HIGH PRIORITY. Speaker driver (GrowBot #9)", "Amazon", 5),
    ("A", "Buy", "Voice/Lights", "Speaker 8 ohm, 0.5-3 W (small)", 1, "HIGH PRIORITY. Audio out (GrowBot #10)", "Amazon", 2),
    ("A", "Buy", "Voice/Lights", "WS2812B LED ring, 7 px, 5 V", 1, "Expressive 'personality' output (GrowBot #11)", "Amazon", 3),
    # -- Power (dual-rail: 5.1V Pi + 6.0V motors) --
    ("A", "Buy", "Power", "18650 Li-ion cell, 2500-3500 mAh (1S)", 1, "Main battery (you don't have one) (GrowBot #12)", "Amazon", 6),
    ("A", "Buy", "Power", "18650 battery holder", 1, "Cell holder + contacts", "Amazon", 2),
    ("A", "Buy", "Power", "TP4056 USB-C charge + protect module", 1, "In-place recharging; PROTECTED OUT+/OUT- version (GrowBot #13)", "Amazon", 2),
    ("A", "Buy", "Power", "MT3608 boost module (x2)", 2,
     "Dual-rail off the 1S cell: ~5.1V (Pi + amp + LED) + ~6.0V (N20 motor rail = TB6612 VM). Isolates Pi from stall brownout (GrowBot #14)", "Amazon", 2),
    ("A", "Buy", "Power", "Electrolytic cap 470-1000 uF, 10 V+", 2, "One per rail; pair with firmware PWM soft-start + drive watchdog (GrowBot #15)", "Amazon", 1),
    ("A", "Buy", "Power", "SPST power switch", 1, "Main cutoff (GrowBot #16)", "Amazon", 2),
    # -- Build / safety --
    ("A", "Buy", "Build", "Hookup wire + Dupont jumpers", 1, "Wire per the GrowBot diagram + the TB6612", "Amazon", 8),
    ("A", "Buy", "Build", "330-470 ohm resistor (WS2812 data line)", 1, "Series resistor on the LED data line (3.3V GPIO -> 5V ring)", "Amazon", 1),
    ("A", "Buy", "Build", "3M double-sided mounting squares", 1, "GrowBot's no-screw board mounting (or print mounts)", "Amazon", 5),
    ("A", "Buy", "Build", "PETG filament (known-safe), 1 spool", 1, "Skip if your printer came with filament / use PLA for the first test print", "Amazon", 25),
    ("A", "Buy", "Safety", "Captive lure (feather/pom on short rigid/braided arm)", 1, "#1 cat-safety item: NO free end, over-molded. Attended play only", "craft / Amazon", 5),
    # -- Print (you have a printer) --
    ("A", "Print", "Drivetrain", "N20 -> 3216 motor-mount adapter", 2, "3216 mounts fit its DC motors, not N20s - print an adapter", "3D print", 0),
    ("A", "Print", "Structure", "Top cover + wheel shroud", 1, "Cat-safety: shroud wheel gaps (<4-6 mm or >25 mm), cover electronics", "3D print", 0),
    ("A", "Print", "Structure", "Captive-lure boom arm", 1, "Rigid arm holding the lure, no free end", "3D print", 0),
    ("A", "Print", "Structure", "Battery + board mounts (optional)", 1, "Or use the 3M squares", "3D print", 0),
    # -- Have (don't buy) --
    ("A", "Have", "Compute", "Raspberry Pi Zero 2 W", 1, "In your Pi kit", "(have)", 0),
    ("A", "Have", "Compute", "2x20 GPIO header", 1, "In your Pi kit - solder it on (you do this at work)", "(have)", 0),
    ("A", "Have", "Compute", "Cables incl. USB power cable", 1, "In your Pi kit", "(have)", 0),
    ("A", "Have", "Chassis", "Adafruit 3216 chassis - frame + caster ball + hardware", 1, "Your base kit = the robot body", "(have)", 0),
    ("A", "Spare", "Drivetrain", "3216 DC motors + 2 wheels", 1, "Superseded by the N20 drivetrain - keep as spares / backup", "(have)", 0),

    # ================= STAGE B — standalone onboard (track Roy's white chest; software-only, $0) =================
    ("B", "Have", "Perception", "Roy's natural white chest patch (no collar, no purchase)", 1,
     "Stage B is SOFTWARE-ONLY on Stage-A hardware: the onboard HSV blob tracker locks onto Roy's white chest (she's black with a clear white chest - a big high-contrast blob). Measure its size to calibrate blob-area -> range. $0, no new parts", "(Roy)", 0),

    # ================= STAGE C — full standalone, no collar (brain upgrade) =================
    ("C", "Buy", "Compute", "Raspberry Pi 5 (4-8 GB)", 1,
     "Brain upgrade for onboard cat detection (Zero 2 W can't). Reuses sensors / drivetrain / power / dock. Iteration 2+", "Amazon", 70),
    ("C", "Buy", "Compute", "AI accelerator - Coral USB TPU or Pi AI HAT (Hailo-8L)", 1,
     "Runs a real on-device cat detector at useful fps, no collar. Pairs with the Pi 5", "Coral / Pimoroni", 65),
    ("C", "Buy", "Power", "Larger battery / >=3A supply", 1,
     "Pi 5 + accelerator draw far more than the Zero 2 W", "Amazon", 15),
]

TOOLS = [
    ("Soldering iron + solder", "Header onto the Pi + wire boards/motors (becqu does this at work)"),
    ("Multimeter", "Set each MT3608 to its target voltage BEFORE connecting the Pi"),
    ("3D printer", "On hand - for the Print rows"),
    ("Digital calipers", "Measure real wheel OD + track for the sim refit"),
    ("Kitchen scale", "Assembled weight for the sim refit"),
    ("Small screwdriver / hex set", "Assembly"),
]

# ---- styling ----
HEAD_FILL = PatternFill("solid", fgColor="1F3A8A")
HEAD_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=15, color="1F3A8A")
THIN = Side(style="thin", color="D0D4DC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(vertical="top", wrap_text=True)
TOP = Alignment(vertical="top")
MONEY = "$#,##0.00"
BUY = PatternFill("solid", fgColor="FBE3E0")
HAVE = PatternFill("solid", fgColor="E3F2E5")
PRINT = PatternFill("solid", fgColor="E1ECF7")
SPARE = PatternFill("solid", fgColor="ECECEC")


def build():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BOM"

    ws.merge_cells("A1:J1")
    ws["A1"] = "Roybot - Staged Bill of Materials"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A2:J2")
    ws["A2"] = ("STAGE A = first prototype, buy NOW (drive + chase brain + senses + power; perception offloaded to your PC/RTX 3070). "
                "STAGE B = standalone onboard, tracks Roy's natural white chest (software-only on A's hardware, no new parts, $0). "
                "STAGE C = OPTIONAL robustness (Pi 5 + AI detector) only if the chest blob is too fragile. B and C build on A. "
                "RED = buy, BLUE = 3D-print, GREEN = have, GREY = spare. Tools on hand. Prices = rough ballparks.")
    ws["A2"].font = Font(italic=True, size=10, color="666666")
    ws["A2"].alignment = WRAP
    ws.row_dimensions[2].height = 54

    headers = ["Stage", "Status", "Group", "Component", "Qty", "Spec / note",
               "Source", "Est. unit (USD)", "Est. line (USD)", "Link / Part # (you fill)"]
    head_row = 9
    first = head_row + 1
    last = first + len(ROWS) - 1
    stage_rng, status_rng, line_rng = f"A{first}:A{last}", f"B{first}:B{last}", f"I{first}:I{last}"

    def summ(row, label, formula, color):
        c = ws.cell(row=row, column=7, value=label); c.font = Font(bold=True); c.alignment = Alignment(horizontal="right")
        v = ws.cell(row=row, column=9, value=formula); v.number_format = MONEY; v.font = Font(bold=True, color=color)

    summ(4, "STAGE A - buy now:", f'=SUMIFS({line_rng},{stage_rng},"A",{status_rng},"Buy")', "B3261E")
    summ(5, "Stage B add-on (later):", f'=SUMIFS({line_rng},{stage_rng},"B",{status_rng},"Buy")', "666666")
    summ(6, "Stage C add-on (later):", f'=SUMIFS({line_rng},{stage_rng},"C",{status_rng},"Buy")', "666666")
    summ(7, "Full project (all buy):", f'=SUMIF({status_rng},"Buy",{line_rng})', "1F3A8A")

    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=head_row, column=c, value=h)
        cell.fill = HEAD_FILL; cell.font = HEAD_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True); cell.border = BORDER

    for i, (stage, status, grp, comp, qty, spec, src, unit) in enumerate(ROWS):
        r = first + i
        vals = [stage, status, grp, comp, qty, spec, src, unit, f"=E{r}*H{r}", ""]
        for c, val in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = BORDER
            cell.alignment = WRAP if c in (4, 6) else TOP
            if c in (8, 9):
                cell.number_format = MONEY

    DataValidation(type="list", formula1='"%s"' % ",".join(STAGES), allow_blank=True)  # (stage dropdown)
    dv_stage = DataValidation(type="list", formula1='"%s"' % ",".join(STAGES), allow_blank=True)
    dv_status = DataValidation(type="list", formula1='"%s"' % ",".join(STATUSES), allow_blank=True)
    ws.add_data_validation(dv_stage); dv_stage.add(stage_rng)
    ws.add_data_validation(dv_status); dv_status.add(status_rng)
    ws.conditional_formatting.add(status_rng, CellIsRule(operator="equal", formula=['"Buy"'], fill=BUY))
    ws.conditional_formatting.add(status_rng, CellIsRule(operator="equal", formula=['"Have"'], fill=HAVE))
    ws.conditional_formatting.add(status_rng, CellIsRule(operator="equal", formula=['"Print"'], fill=PRINT))
    ws.conditional_formatting.add(status_rng, CellIsRule(operator="equal", formula=['"Spare"'], fill=SPARE))

    ws.auto_filter.ref = f"A{head_row}:J{last}"
    for i, w in enumerate([7, 9, 13, 40, 5, 48, 20, 13, 13, 20], start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.freeze_panes = f"A{first}"

    ts = wb.create_sheet("Tools")
    ts.merge_cells("A1:B1"); ts["A1"] = "Tools (assumed on hand)"; ts["A1"].font = TITLE_FONT
    for c, h in enumerate(["Tool", "Needed for"], start=1):
        cell = ts.cell(row=3, column=c, value=h); cell.fill = HEAD_FILL; cell.font = HEAD_FONT; cell.border = BORDER
    for i, (tool, use) in enumerate(TOOLS, start=4):
        ts.cell(row=i, column=1, value=tool).border = BORDER
        a = ts.cell(row=i, column=2, value=use); a.border = BORDER; a.alignment = WRAP
    ts.column_dimensions["A"].width = 28; ts.column_dimensions["B"].width = 64; ts.freeze_panes = "A4"

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    na = sum(1 for r in ROWS if r[0] == "A" and r[1] == "Buy")
    print(f"wrote {OUT}  ({len(ROWS)} rows; Stage A buy = {na} items)")


if __name__ == "__main__":
    build()
